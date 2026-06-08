# app_odp/services/dashboard_service.py

from datetime import date, datetime, timedelta
import re
import json
from flask import request
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import func, and_
from app_odp.policy.policy import RbacPolicy
from app_odp.models import (
    db,
    InputOdp,
    InputOdpLog,
    OdpRuntimeLog,
    ProductionKpiSnapshot,
    ProductionCapacityCalendar,
    User,
    Reparti,
    Risorse,
    Lavorazioni,
)

from app_odp.routes import (
    ROME_TZ,
    _active_value_for_phase,
    _base_odp_query,
    _fase_to_int,
    _first_code_from_cell,
    _norm_text,
    _ordine_ref_label,
    _safe_float,
    _parse_phase_list,
)

DASHBOARD_FILTER_KEYS = (
    "reparto",
    "risorsa",
    "lavorazione",
    "operatore",
    "articolo",
    "stato",
)

DASHBOARD_PRODUZIONE_FUTURE_DAYS = 31


def _dashboard_is_chiusa(ordine: InputOdp) -> bool:
    stato = _dashboard_stato_norm(ordine).lower()
    return "chius" in stato or "terminat" in stato


def _dashboard_reparto_label(
    codice: str, label_map: dict[str, str] | None = None
) -> str:
    codice = _first_code_from_cell(codice)
    if not codice:
        return "-"

    label_map = label_map if label_map is not None else _dashboard_reparti_label_map()
    return label_map.get(codice, codice)


def _dashboard_lavorazione_attiva(ordine: InputOdp) -> str:
    runtime = getattr(ordine, "runtime_row", None)

    return (
        _norm_text(getattr(runtime, "LavorazioneAttiva", ""))
        or _norm_text(getattr(ordine, "LavorazioneAttiva", ""))
        or _first_code_from_cell(_dashboard_active_value(ordine, "CodLavorazione"))
    )


def _dashboard_attrezzaggio_ore(ordine: InputOdp) -> float:
    raw = getattr(ordine, "AttrezzaggioAttivo", "") or getattr(
        ordine, "TempoAttrezzaggio", ""
    )

    value = _safe_float(raw)

    # Attrezzaggio nel tuo codice storico era gestito come minuti.
    return value / 60.0 if value > 0 else 0.0


def _dashboard_carico_ore(ordine: InputOdp) -> float:
    return round(
        _dashboard_tempo_previsto_ore(ordine) + _dashboard_attrezzaggio_ore(ordine),
        2,
    )


def _dashboard_order_label(ordine: InputOdp) -> str:
    try:
        return _ordine_ref_label(ordine)
    except Exception:
        return _norm_text(getattr(ordine, "NumProgrRiga", "")) or _norm_text(
            getattr(ordine, "IdDocumento", "")
        )


def _dashboard_capacity_by_weekday(
    *,
    scope_type: str = "global",
    scope_code: str = "*",
    fallback_to_global: bool = True,
) -> dict[int, float]:
    rows = (
        ProductionCapacityCalendar.query.filter(
            ProductionCapacityCalendar.active.is_(True)
        )
        .filter(ProductionCapacityCalendar.scope_type == scope_type)
        .filter(ProductionCapacityCalendar.scope_code == scope_code)
        .all()
    )

    out = {i: 0.0 for i in range(7)}

    for row in rows:
        out[int(row.weekday)] = float(row.hours_capacity or 0.0)

    if fallback_to_global and scope_type != "global" and not rows:
        return _dashboard_capacity_by_weekday(
            scope_type="global",
            scope_code="*",
        )

    return out


def _dashboard_carico_prossimo_mese(
    ordini: list[InputOdp],
    *,
    capacity_by_weekday: dict[int, float] | None = None,
    capacity_operator_count: int = 0,
) -> list[dict]:
    capacity = capacity_by_weekday or _dashboard_capacity_by_weekday()

    by_day = {
        day.isoformat(): {
            "date": day.isoformat(),
            "label": day.strftime("%d/%m"),
            "ore_arretrate": 0.0,
            "ore_senza_scadenza": 0.0,
            "ore_pianificate": 0.0,
            "ore_attive": 0.0,
            "ore_sospese": 0.0,
            "capacita": float(capacity.get(day.weekday(), 0.0) or 0.0),
            "operatori_capacita": int(capacity_operator_count or 0),
        }
        for day in _dashboard_next_month_days()
    }

    today = _dashboard_today()
    end_day = today + timedelta(days=DASHBOARD_PRODUZIONE_FUTURE_DAYS - 1)
    today_key = today.isoformat()

    for ordine in ordini:
        stato = _dashboard_stato_norm(ordine).lower()
        data_fine = _dashboard_data_fine_prevista(ordine)
        ore = _dashboard_carico_ore(ordine)

        if not ore:
            continue

        if not data_fine:
            # Gli ordini aperti senza data schedulata sono comunque carico reale:
            # li evidenziamo sul primo giorno del cruscotto.
            by_day[today_key]["ore_senza_scadenza"] += ore
            continue

        if data_fine < today:
            # Gli ordini aperti già scaduti non devono sparire dal grafico:
            # sono arretrato da gestire subito.
            by_day[today_key]["ore_arretrate"] += ore
            continue

        if data_fine > end_day:
            continue

        day_key = data_fine.isoformat()
        if day_key not in by_day:
            continue

        if "pianificat" in stato:
            by_day[day_key]["ore_pianificate"] += ore
        elif "attiv" in stato:
            by_day[day_key]["ore_attive"] += ore
        elif "sospes" in stato:
            by_day[day_key]["ore_sospese"] += ore

    for row in by_day.values():
        row["ore_arretrate"] = round(row["ore_arretrate"], 2)
        row["ore_senza_scadenza"] = round(row["ore_senza_scadenza"], 2)
        row["ore_pianificate"] = round(row["ore_pianificate"], 2)
        row["ore_attive"] = round(row["ore_attive"], 2)
        row["ore_sospese"] = round(row["ore_sospese"], 2)
        row["ore_totali"] = round(
            row["ore_arretrate"]
            + row["ore_senza_scadenza"]
            + row["ore_pianificate"]
            + row["ore_attive"]
            + row["ore_sospese"],
            2,
        )
        row["sovraccarico"] = bool(
            row["capacita"] > 0 and row["ore_totali"] > row["capacita"]
        )

    return list(by_day.values())


def _dashboard_capacity_hours_from_weekday(
    capacity_by_weekday: dict[int, float],
    *,
    days: list[date] | None = None,
) -> float:
    """Somma le ore di capacità sulla finestra del cruscotto, da oggi in avanti."""
    days = days or _dashboard_next_month_days()

    return round(
        sum(float(capacity_by_weekday.get(day.weekday(), 0.0) or 0.0) for day in days),
        2,
    )


def _dashboard_ordine_in_saturation_window(ordine: InputOdp) -> bool:
    """
    Include nel carico di saturazione:
    - ordini senza data fine prevista;
    - ordini arretrati;
    - ordini con fine prevista entro la finestra del cruscotto.
    """
    data_fine = _dashboard_data_fine_prevista(ordine)

    if data_fine is None:
        return True

    today = _dashboard_today()
    end_day = today + timedelta(days=DASHBOARD_PRODUZIONE_FUTURE_DAYS - 1)

    return data_fine <= end_day


def _dashboard_capacity_by_reparto_for_policy(
    policy: RbacPolicy,
    filters: dict | None = None,
) -> dict[str, dict]:
    """
    Calcola la capacità disponibile per reparto sommando le ore degli utenti
    attivi appartenenti al reparto, sulla finestra da oggi in avanti.
    """
    buckets: dict[str, dict] = {}
    reparto_labels = _dashboard_reparti_label_map()

    for user in _dashboard_capacity_users(policy, filters):
        reparto = _norm_text(getattr(user, "RepartoPrinc", ""))

        if not reparto:
            continue

        capacity = _dashboard_capacity_for_operator(user)
        ore_disponibili = _dashboard_capacity_hours_from_weekday(capacity)

        if ore_disponibili <= 0:
            continue

        reparto_label = _dashboard_reparto_label(reparto, reparto_labels)

        buckets.setdefault(
            reparto,
            {
                "label": reparto_label,
                "reparto": reparto,
                "codice_reparto": reparto,
                "capacita": 0.0,
                "operatori_capacita": 0,
            },
        )

        buckets[reparto]["capacita"] += ore_disponibili
        buckets[reparto]["operatori_capacita"] += 1

    for row in buckets.values():
        row["capacita"] = round(float(row["capacita"] or 0.0), 2)

    return buckets


def _dashboard_capacity_hours_today(capacity_by_weekday: dict[int, float]) -> float:
    """
    Restituisce la capacità dell'operatore/reparto per il solo giorno attuale.

    Esempio:
    - oggi lunedì -> usa weekday 0;
    - oggi venerdì -> usa weekday 4.
    """
    today = _dashboard_today()
    return round(float(capacity_by_weekday.get(today.weekday(), 0.0) or 0.0), 2)


def _dashboard_ordine_in_saturation_day(ordine: InputOdp) -> bool:
    """
    Stabilisce se l'ordine deve pesare nella saturazione di oggi.

    Regola consigliata:
    - ordini senza data fine prevista: inclusi oggi;
    - ordini arretrati: inclusi oggi;
    - ordini con data fine prevista oggi: inclusi oggi;
    - ordini futuri: esclusi.
    """
    data_fine = _dashboard_data_fine_prevista(ordine)

    if data_fine is None:
        return True

    today = _dashboard_today()

    return data_fine <= today


def _dashboard_capacity_by_reparto_today_for_policy(
    policy: RbacPolicy,
    filters: dict | None = None,
) -> dict[str, dict]:
    """
    Calcola la capacità disponibile oggi per reparto.

    La capacità reparto è la somma delle ore disponibili oggi
    degli operatori attivi appartenenti a quel reparto.
    """
    buckets: dict[str, dict] = {}
    reparto_labels = _dashboard_reparti_label_map()

    for user in _dashboard_capacity_users(policy, filters):
        reparto = _norm_text(getattr(user, "RepartoPrinc", ""))

        if not reparto:
            continue

        capacity = _dashboard_capacity_for_operator(user)
        ore_oggi = _dashboard_capacity_hours_today(capacity)

        if ore_oggi <= 0:
            continue

        reparto_label = _dashboard_reparto_label(reparto, reparto_labels)

        buckets.setdefault(
            reparto,
            {
                "label": reparto_label,
                "reparto": reparto,
                "codice_reparto": reparto,
                "capacita": 0.0,
                "operatori_capacita": 0,
            },
        )

        buckets[reparto]["capacita"] += ore_oggi
        buckets[reparto]["operatori_capacita"] += 1

    for row in buckets.values():
        row["capacita"] = round(float(row["capacita"] or 0.0), 2)

    return buckets


def _dashboard_saturazione_risorse(
    ordini: list[InputOdp],
    policy: RbacPolicy,
    filters: dict | None = None,
) -> list[dict]:
    """
    Mantiene il nome storico `saturazione_risorse` per non rompere il frontend,
    ma calcola la saturazione per reparto sul solo giorno attuale.

    Formula:
        ore ordini reparto oggi / ore disponibili oggi operatori reparto * 100
    """
    reparto_labels = _dashboard_reparti_label_map()
    capacity_by_reparto = _dashboard_capacity_by_reparto_today_for_policy(
        policy,
        filters,
    )

    buckets: dict[str, dict] = {}
    seen_keys = set()

    for ordine in ordini or []:
        stato = _dashboard_stato_norm(ordine).lower()

        if not ("attiv" in stato or "sospes" in stato or "pianificat" in stato):
            continue

        if not _dashboard_ordine_in_saturation_day(ordine):
            continue

        ordine_key = _dashboard_order_key(ordine)

        if ordine_key in seen_keys:
            continue

        seen_keys.add(ordine_key)

        reparto = _dashboard_reparto_attivo(ordine) or "-"
        reparto_label = _dashboard_reparto_label(reparto, reparto_labels)
        ore = _dashboard_carico_ore(ordine)

        buckets.setdefault(
            reparto,
            {
                "label": reparto_label,
                "reparto": reparto,
                "codice_reparto": reparto,
                "ore_attive": 0.0,
                "ore_sospese": 0.0,
                "ore_pianificate": 0.0,
                "ore_totali": 0.0,
            },
        )

        if "attiv" in stato:
            buckets[reparto]["ore_attive"] += ore
        elif "sospes" in stato:
            buckets[reparto]["ore_sospese"] += ore
        elif "pianificat" in stato:
            buckets[reparto]["ore_pianificate"] += ore

    # Mostra anche reparti con capacità oggi ma senza carico.
    for reparto, cap_row in capacity_by_reparto.items():
        buckets.setdefault(
            reparto,
            {
                "label": cap_row["label"],
                "reparto": reparto,
                "codice_reparto": reparto,
                "ore_attive": 0.0,
                "ore_sospese": 0.0,
                "ore_pianificate": 0.0,
                "ore_totali": 0.0,
            },
        )

    out = []

    for reparto, row in buckets.items():
        cap_row = capacity_by_reparto.get(reparto) or {}

        ore_attive = round(float(row.get("ore_attive") or 0.0), 2)
        ore_sospese = round(float(row.get("ore_sospese") or 0.0), 2)
        ore_pianificate = round(float(row.get("ore_pianificate") or 0.0), 2)
        ore_totali = round(ore_attive + ore_sospese + ore_pianificate, 2)
        capacita = round(float(cap_row.get("capacita") or 0.0), 2)

        if capacita > 0:
            saturazione = round((ore_totali / capacita) * 100, 2)
        elif ore_totali > 0:
            saturazione = 120.0
        else:
            saturazione = 0.0

        if capacita <= 0 and ore_totali > 0:
            livello = "senza_capacita"
        elif saturazione > 100:
            livello = "critico"
        elif saturazione >= 80:
            livello = "attenzione"
        elif saturazione >= 50:
            livello = "buono"
        else:
            livello = "sottocarico"

        out.append(
            {
                "label": row.get("label") or reparto,
                "reparto": reparto,
                "codice_reparto": reparto,
                "ore_attive": ore_attive,
                "ore_sospese": ore_sospese,
                "ore_pianificate": ore_pianificate,
                "ore_totali": ore_totali,
                "capacita": capacita,
                "scostamento_ore": round(capacita - ore_totali, 2),
                "saturazione": saturazione,
                "livello": livello,
                "operatori_capacita": int(cap_row.get("operatori_capacita") or 0),
            }
        )

    return sorted(out, key=lambda x: (-x["saturazione"], x["label"].lower()))[:12]


def _dashboard_ordini_per_reparto(ordini: list[InputOdp]) -> list[dict]:
    buckets = {}
    reparto_labels = _dashboard_reparti_label_map()
    seen_keys = set()

    for ordine in ordini or []:
        stato = _dashboard_stato_norm(ordine).lower()

        if not ("attiv" in stato or "sospes" in stato or "pianificat" in stato):
            continue

        ordine_key = (
            _norm_text(getattr(ordine, "IdDocumento", "")),
            _norm_text(getattr(ordine, "IdRiga", "")),
            _norm_text(getattr(ordine, "FaseAttiva", "")) or "1",
        )

        if ordine_key in seen_keys:
            continue

        seen_keys.add(ordine_key)

        reparto = _dashboard_reparto_attivo(ordine) or "-"
        reparto_label = _dashboard_reparto_label(reparto, reparto_labels)

        buckets.setdefault(
            reparto,
            {
                "label": reparto_label,
                "reparto": reparto,
                "codice_reparto": reparto,
                "ordini_attivi": 0,
                "ordini_sospesi": 0,
                "ordini_pianificati": 0,
                "ordini_totali": 0,
            },
        )

        if "attiv" in stato:
            buckets[reparto]["ordini_attivi"] += 1
        elif "sospes" in stato:
            buckets[reparto]["ordini_sospesi"] += 1
        elif "pianificat" in stato:
            buckets[reparto]["ordini_pianificati"] += 1

    out = []

    for row in buckets.values():
        row["ordini_totali"] = (
            int(row["ordini_attivi"] or 0)
            + int(row["ordini_sospesi"] or 0)
            + int(row["ordini_pianificati"] or 0)
        )
        out.append(row)

    return sorted(
        out,
        key=lambda x: (-int(x["ordini_totali"] or 0), x["label"].lower()),
    )[:12]


def _dashboard_ordini_per_risorsa_chart(carico_rows: list[dict]) -> list[dict]:
    out = []

    for row in carico_rows or []:
        ordini_attivi = int(row.get("ordini_attivi") or 0)
        ordini_sospesi = int(row.get("ordini_sospesi") or 0)
        ordini_pianificati = int(row.get("ordini_pianificati") or 0)

        out.append(
            {
                "label": row.get("risorsa") or "-",
                "risorsa": row.get("risorsa") or "-",
                "ordini_attivi": ordini_attivi,
                "ordini_sospesi": ordini_sospesi,
                "ordini_pianificati": ordini_pianificati,
                "ordini_totali": ordini_attivi + ordini_sospesi + ordini_pianificati,
            }
        )

    return sorted(
        out,
        key=lambda x: (-int(x["ordini_totali"] or 0), x["label"].lower()),
    )[:12]


def _dashboard_is_collaudo(ordine: InputOdp) -> bool:
    reparto = _dashboard_reparto_attivo(ordine)
    risorsa = _dashboard_risorsa_attiva(ordine).lower()
    lavorazione = _dashboard_lavorazione_attiva(ordine).lower()

    return reparto == "70" or "coll" in risorsa or "coll" in lavorazione


def _dashboard_user_matches_capacity_filters(user: User, filters: dict | None) -> bool:
    filters = filters or {}

    operatore_filter = _dashboard_text_filter(filters.get("operatore"))
    reparto_filter = _dashboard_text_filter(filters.get("reparto"))
    risorsa_filter = _dashboard_text_filter(filters.get("risorsa"))

    username = _norm_text(getattr(user, "username", ""))
    reparto_code = _norm_text(getattr(user, "RepartoPrinc", ""))

    if operatore_filter and operatore_filter not in username.lower():
        return False

    if reparto_filter:
        reparto_label = reparto_code

        if reparto_code:
            reparto = Reparti.query.filter(
                func.lower(Reparti.Codice) == reparto_code.lower()
            ).first()

            if reparto is not None:
                reparto_label = (
                    f"{_norm_text(getattr(reparto, 'Codice', ''))} "
                    f"{_norm_text(getattr(reparto, 'Descrizione', ''))}"
                )

        if reparto_filter not in reparto_label.lower():
            return False

    if risorsa_filter:
        risorsa_labels = []

        for risorsa in getattr(user, "risorse", []) or []:
            risorsa_labels.append(
                f"{_norm_text(getattr(risorsa, 'Codice', ''))} "
                f"{_norm_text(getattr(risorsa, 'Descrizione', ''))}"
            )

        if not any(risorsa_filter in label.lower() for label in risorsa_labels):
            return False

    return True


def _dashboard_capacity_for_operator(user: User) -> dict[int, float]:
    """
    Capacità settimanale del singolo operatore.

    Priorità:
    1. capacità specifica operatore:
       scope_type = "operatore", scope_code = User.id

    2. capacità reparto:
       scope_type = "reparto", scope_code = User.RepartoPrinc

    3. capacità globale:
       scope_type = "global", scope_code = "*"

    Nota:
    questa funzione restituisce sempre la settimana completa.
    La saturazione reparto usa poi solo il weekday di oggi.
    """
    if user is None:
        return {i: 0.0 for i in range(7)}

    operator_code = str(int(user.id))

    if _dashboard_capacity_rows_exist("operatore", operator_code):
        return _dashboard_capacity_by_weekday(
            scope_type="operatore",
            scope_code=operator_code,
            fallback_to_global=False,
        )

    reparto_code = _norm_text(getattr(user, "RepartoPrinc", ""))

    if reparto_code and _dashboard_capacity_rows_exist("reparto", reparto_code):
        return _dashboard_capacity_by_weekday(
            scope_type="reparto",
            scope_code=reparto_code,
            fallback_to_global=False,
        )

    if _dashboard_capacity_rows_exist("global", "*"):
        return _dashboard_capacity_by_weekday(
            scope_type="global",
            scope_code="*",
            fallback_to_global=False,
        )

    return {i: 0.0 for i in range(7)}


def _dashboard_seed_user_filter_options(options: dict) -> None:
    """
    Alimenta il filtro Operatore da tutti gli utenti presenti in users.
    Non filtra per active, ruolo, reparto o policy.
    """
    users = User.query.order_by(func.lower(User.username)).all()

    for user in users:
        username = _norm_text(getattr(user, "username", ""))

        if not username:
            continue

        _dashboard_add_filter_option(
            options,
            "operatore",
            username,
            username,
        )


def _dashboard_empty_filter_options() -> dict:
    return {key: [] for key in DASHBOARD_FILTER_KEYS}


def _dashboard_new_filter_options_bucket() -> dict:
    return {key: {} for key in DASHBOARD_FILTER_KEYS}


def _dashboard_filter_label(value: str, description: str = "") -> str:
    value = _norm_text(value)
    description = _norm_text(description)

    if not value:
        return ""

    if description and description.lower() != value.lower():
        return f"{description} ({value})"

    return value


def _dashboard_model_label_map(model) -> dict[str, str]:
    rows = model.query.order_by(
        func.lower(func.coalesce(model.Descrizione, model.Codice)),
        func.lower(model.Codice),
    ).all()

    out = {}

    for row in rows:
        codice = _norm_text(getattr(row, "Codice", ""))
        descrizione = _norm_text(getattr(row, "Descrizione", ""))

        if codice:
            out[codice] = _dashboard_filter_label(codice, descrizione)

    return out


def _dashboard_filter_label_maps() -> dict:
    return {
        "reparto": _dashboard_model_label_map(Reparti),
        "risorsa": _dashboard_model_label_map(Risorse),
        "lavorazione": _dashboard_model_label_map(Lavorazioni),
    }


def _dashboard_add_filter_option(
    options: dict,
    key: str,
    value,
    label: str = "",
) -> None:
    if key not in options:
        return

    value = _norm_text(value)
    if not value:
        return

    options[key].setdefault(
        value,
        {
            "value": value,
            "label": _norm_text(label) or value,
        },
    )


def _dashboard_article_filter_label(cod_art: str, descrizione: str = "") -> str:
    cod_art = _norm_text(cod_art)
    descrizione = _norm_text(descrizione)

    if cod_art and descrizione:
        return f"{cod_art} - {descrizione}"

    return cod_art or descrizione


def _dashboard_collect_filter_options_from_row(
    options: dict,
    row: dict,
    label_maps: dict,
) -> None:
    articolo = _norm_text(row.get("cod_art") or row.get("articolo"))

    _dashboard_add_filter_option(
        options,
        "articolo",
        articolo,
        _dashboard_article_filter_label(articolo, row.get("descrizione")),
    )


def _dashboard_finalize_filter_options(options: dict) -> dict:
    out = {}

    for key in DASHBOARD_FILTER_KEYS:
        rows = list((options.get(key) or {}).values())

        rows.sort(
            key=lambda row: (
                _norm_text(row.get("label")).lower(),
                _norm_text(row.get("value")).lower(),
            )
        )

        out[key] = rows

    return out


def _dashboard_text_filter(value) -> str:
    return _norm_text(value).lower()


def _dashboard_cruscotto_filters_from_request() -> dict:
    return {
        "reparto": _dashboard_text_filter(request.args.get("reparto")),
        "risorsa": _dashboard_text_filter(request.args.get("risorsa")),
        "lavorazione": _dashboard_text_filter(request.args.get("lavorazione")),
        "operatore": _dashboard_text_filter(request.args.get("operatore")),
        "articolo": _dashboard_text_filter(request.args.get("articolo")),
        "stato": _dashboard_text_filter(request.args.get("stato")),
    }


def _dashboard_row_matches_filters(row: dict, filters: dict) -> bool:
    for key in ("reparto", "risorsa", "lavorazione", "operatore", "articolo", "stato"):
        expected = _dashboard_text_filter(filters.get(key))
        if not expected:
            continue

        if key == "articolo":
            current = " ".join(
                [
                    _norm_text(row.get("cod_art")),
                    _norm_text(row.get("descrizione")),
                ]
            ).lower()
        else:
            current = _norm_text(row.get(key)).lower()

        if expected not in current:
            return False

    return True


def _capacity_calendar_payload() -> list[dict]:
    rows = (
        ProductionCapacityCalendar.query.filter(
            ProductionCapacityCalendar.active.is_(True)
        )
        .order_by(
            ProductionCapacityCalendar.scope_type.asc(),
            ProductionCapacityCalendar.scope_code.asc(),
            ProductionCapacityCalendar.weekday.asc(),
        )
        .all()
    )

    return [
        {
            "id": row.id,
            "scope_type": row.scope_type,
            "scope_code": row.scope_code,
            "weekday": row.weekday,
            "hours_capacity": float(row.hours_capacity or 0),
        }
        for row in rows
    ]


def _dashboard_produzione_allowed_sections(policy: RbacPolicy) -> dict:
    return {
        "cruscotto": bool(policy.can("dashboard_produzione")),
        "kpi": bool(policy.can("kpi_produzione")),
        "export": bool(policy.can("kpi_export")),
        "config": bool(policy.can("kpi_config")),
    }


def _dashboard_produzione_initial_payload(policy: RbacPolicy) -> dict:
    return {
        "allowed_sections": _dashboard_produzione_allowed_sections(policy),
        "capacity_calendar": _capacity_calendar_payload(),
        "cruscotto": {
            "cards": {},
            "charts": {},
            "criticita": [],
            "filter_options": _dashboard_empty_filter_options(),
        },
        "kpi": {
            "cards": {},
            "charts": {},
            "details": [],
            "filter_options": _dashboard_empty_filter_options(),
        },
    }


def _dashboard_seed_model_filter_options(
    options: dict,
    *,
    key: str,
    model,
) -> None:
    """
    Alimenta un filtro da una tabella anagrafica con Codice/Descrizione.
    Esempi:
    - reparto -> Reparti
    - risorsa -> Risorse
    - lavorazione -> Lavorazioni
    """
    rows = model.query.order_by(
        func.lower(func.coalesce(model.Descrizione, model.Codice)),
        func.lower(model.Codice),
    ).all()

    for row in rows:
        codice = _norm_text(getattr(row, "Codice", ""))
        descrizione = _norm_text(getattr(row, "Descrizione", ""))

        if not codice:
            continue

        _dashboard_add_filter_option(
            options,
            key,
            codice,
            _dashboard_filter_label(codice, descrizione),
        )


def _dashboard_today() -> date:
    return datetime.now(ROME_TZ).date()


def _dashboard_seed_stato_filter_options(options: dict) -> None:
    """
    Stati standard usati dalla dashboard produzione.
    Non vengono presi da TipologieStato perché quel model contiene solo 'tipo'
    numerico e non la descrizione testuale usata nei dati dashboard.
    """
    for stato in ("Pianificata", "Attivo", "In Sospeso", "Chiusa"):
        _dashboard_add_filter_option(
            options,
            "stato",
            stato,
            stato,
        )


def _dashboard_build_cruscotto_payload(policy: RbacPolicy) -> dict:
    payload = _dashboard_cruscotto_empty_payload()
    filters = _dashboard_cruscotto_filters_from_request()
    payload["filters"] = filters

    ordini = list(policy.filter_input_odp(_base_odp_query()).all())

    today = _dashboard_today()

    operatori = {}
    carico_risorsa = {}
    criticita = []
    collaudo_rows = []
    filtered_ordini = []
    filter_options = _dashboard_new_filter_options_bucket()
    filter_label_maps = _dashboard_filter_label_maps()
    _dashboard_seed_master_filter_options(filter_options)

    stati_chart = {
        "Pianificata": 0,
        "Attivo": 0,
        "In Sospeso": 0,
    }

    for ordine in ordini:
        if _dashboard_is_chiusa(ordine):
            continue

        stato_raw = _dashboard_stato_norm(ordine)
        stato = stato_raw.lower()
        ore = _dashboard_carico_ore(ordine)
        data_fine = _dashboard_data_fine_prevista(ordine)
        runtime = getattr(ordine, "runtime_row", None)
        base_row = _dashboard_order_payload(ordine)
        _dashboard_collect_filter_options_from_row(
            filter_options,
            base_row,
            filter_label_maps,
        )

        if not _dashboard_row_matches_filters(base_row, filters):
            continue

        filtered_ordini.append(ordine)

        is_attivo = "attiv" in stato
        is_sospeso = "sospes" in stato
        is_pianificata = "pianificat" in stato

        if not (is_attivo or is_sospeso or is_pianificata):
            continue
        payload["details"].append(base_row)

        if is_attivo:
            payload["cards"]["ordini_attivi"] += 1
            payload["cards"]["tempo_previsto_residuo"] += ore
            stati_chart["Attivo"] += 1

        elif is_sospeso:
            payload["cards"]["ordini_sospesi"] += 1
            stati_chart["In Sospeso"] += 1

        elif is_pianificata:
            payload["cards"]["ordini_pianificati"] += 1
            stati_chart["Pianificata"] += 1

        if ore <= 0:
            payload["cards"]["ordini_senza_tempo_previsto"] += 1
            criticita.append(_dashboard_order_payload(ordine, "Senza tempo previsto"))

        if data_fine:
            if data_fine == today:
                payload["cards"]["ordini_scadenza_oggi"] += 1
                criticita.append(_dashboard_order_payload(ordine, "Scade oggi"))

            elif data_fine < today:
                payload["cards"]["ordini_in_ritardo"] += 1
                criticita.append(_dashboard_order_payload(ordine, "In ritardo"))

        if is_attivo and runtime is not None:
            operatore = _norm_text(getattr(runtime, "Utente_operazione", ""))
            if operatore:
                operatori.setdefault(
                    operatore,
                    {
                        "operatore": operatore,
                        "ordini_attivi": 0,
                        "ore_attive": 0.0,
                        "ordini": [],
                    },
                )
                operatori[operatore]["ordini_attivi"] += 1
                operatori[operatore]["ore_attive"] += ore
                operatori[operatore]["ordini"].append(_dashboard_order_payload(ordine))

        risorsa = _dashboard_risorsa_attiva(ordine) or "-"
        carico_risorsa.setdefault(
            risorsa,
            {
                "risorsa": risorsa,
                "ordini_attivi": 0,
                "ordini_sospesi": 0,
                "ordini_pianificati": 0,
                "ore_attive": 0.0,
                "ore_sospese": 0.0,
                "ore_pianificate": 0.0,
            },
        )

        if is_attivo:
            carico_risorsa[risorsa]["ordini_attivi"] += 1
            carico_risorsa[risorsa]["ore_attive"] += ore
        elif is_sospeso:
            carico_risorsa[risorsa]["ordini_sospesi"] += 1
            carico_risorsa[risorsa]["ore_sospese"] += ore
        elif is_pianificata:
            carico_risorsa[risorsa]["ordini_pianificati"] += 1
            carico_risorsa[risorsa]["ore_pianificate"] += ore

        if _dashboard_is_collaudo(ordine):
            payload["cards"]["ordini_collaudo"] += 1
            collaudo_rows.append(_dashboard_order_payload(ordine, "Collaudo"))

        # Criticità: attivo da troppo tempo rispetto al previsto.
        if is_attivo and runtime is not None:
            last_activation = getattr(runtime, "data_ultima_attivazione", None)
            started_at = None

            if last_activation:
                try:
                    started_at = datetime.fromisoformat(
                        str(last_activation)
                    ).astimezone(ROME_TZ)
                except Exception:
                    started_at = None

            if started_at and ore > 0:
                elapsed_hours = (
                    datetime.now(ROME_TZ) - started_at
                ).total_seconds() / 3600.0

                if elapsed_hours > ore:
                    record = _dashboard_order_payload(ordine, "Attivo oltre previsto")
                    record["ore_attive_effettive"] = round(elapsed_hours, 2)
                    criticita.append(record)

    payload["cards"]["tempo_previsto_residuo"] = round(
        payload["cards"]["tempo_previsto_residuo"],
        2,
    )
    payload["cards"]["operatori_impegnati"] = len(operatori)

    payload["charts"]["stati_ordine"] = [
        {"label": key, "value": value} for key, value in stati_chart.items()
    ]

    capacity_by_weekday, capacity_operator_count = (
        _dashboard_capacity_by_weekday_for_policy(
            policy,
            filters,
        )
    )

    payload["cards"]["operatori_capacita"] = capacity_operator_count

    payload["charts"]["carico_prossimi_giorni"] = _dashboard_carico_prossimo_mese(
        filtered_ordini,
        capacity_by_weekday=capacity_by_weekday,
        capacity_operator_count=capacity_operator_count,
    )

    payload["operatori"] = sorted(
        [
            {
                **row,
                "ore_attive": round(float(row["ore_attive"] or 0.0), 2),
            }
            for row in operatori.values()
        ],
        key=lambda x: (-x["ordini_attivi"], x["operatore"].lower()),
    )

    payload["carico_risorsa"] = sorted(
        [
            {
                **row,
                "ore_attive": round(float(row["ore_attive"] or 0.0), 2),
                "ore_sospese": round(float(row["ore_sospese"] or 0.0), 2),
                "ore_pianificate": round(float(row["ore_pianificate"] or 0.0), 2),
                "ore_totali": round(
                    float(row["ore_attive"] or 0.0)
                    + float(row["ore_sospese"] or 0.0)
                    + float(row["ore_pianificate"] or 0.0),
                    2,
                ),
            }
            for row in carico_risorsa.values()
        ],
        key=lambda x: (-x["ore_totali"], x["risorsa"].lower()),
    )

    payload["charts"]["carico_per_risorsa"] = _dashboard_carico_per_risorsa_chart(
        payload["carico_risorsa"]
    )

    payload["charts"]["carico_per_reparto"] = _dashboard_carico_per_reparto(
        filtered_ordini
    )

    payload["charts"]["ordini_per_reparto"] = _dashboard_ordini_per_reparto(
        filtered_ordini
    )

    payload["charts"]["ordini_per_risorsa"] = _dashboard_ordini_per_risorsa_chart(
        payload["carico_risorsa"]
    )

    payload["charts"]["saturazione_risorse"] = _dashboard_saturazione_risorse(
        filtered_ordini,
        policy,
        filters,
    )

    payload["cards"]["reparti_sovraccarichi"] = sum(
        1
        for row in payload["charts"]["saturazione_risorse"]
        if row.get("livello") in {"critico", "senza_capacita"}
    )

    # Compatibilità con il frontend attuale:
    # il template legge ancora risorse_sovraccariche.
    payload["cards"]["risorse_sovraccariche"] = payload["cards"][
        "reparti_sovraccarichi"
    ]

    payload["criticita"] = sorted(
        criticita,
        key=lambda x: (
            0 if x["tipo"] == "In ritardo" else 1,
            -int(x.get("ritardo_giorni") or 0),
            x.get("data_fine_prevista") or "9999-12-31",
        ),
    )[:100]
    payload["cards"]["ordini_critici"] = len(payload["criticita"])
    payload["filter_options"] = _dashboard_finalize_filter_options(filter_options)

    payload["collaudo"] = collaudo_rows[:100]
    payload["capacity_calendar"] = _capacity_calendar_payload()
    payload["details"] = sorted(
        payload["details"],
        key=lambda x: (
            x.get("data_fine_prevista") or "9999-12-31",
            x.get("stato") or "",
            x.get("ordine") or "",
        ),
    )[:500]

    return payload


def _dashboard_kpi_empty_payload() -> dict:
    return {
        "cards": {
            "ordini_chiusi": 0,
            "ordini_in_ritardo": 0,
            "percentuale_ritardo": 0.0,
            "giorni_medi_ritardo": 0.0,
            "tempo_previsto_totale": 0.0,
            "tempo_reale_totale": 0.0,
            "scostamento_totale": 0.0,
            "scostamento_medio": 0.0,
            "tempo_medio_ordine": 0.0,
            "tempo_medio_fase": 0.0,
            "tempo_medio_collaudo": 0.0,
            "collaudi_chiusi_oggi": 0,
            "affidabilita_tempi": 0.0,
            "ordini_tempi_affidabili": 0,
            "ordini_tempi_non_affidabili": 0,
        },
        "charts": {
            "tempo_reale_vs_previsto": [],
            "ritardo_medio_reparto": [],
            "ritardo_medio_risorsa": [],
            "top_risorse_scostamento": [],
            "top_reparti_scostamento": [],
            "top_lavorazioni_scostamento": [],
            "affidabilita_tempi": [],
        },
        "details": [],
        "aggregati": {
            "reparti": [],
            "risorse": [],
            "lavorazioni": [],
            "operatori": [],
            "articoli": [],
        },
        "filter_options": _dashboard_empty_filter_options(),
    }


def _kpi_macchine_prodotte(rt: OdpRuntimeLog, il: InputOdpLog | None) -> float:
    """
    Numero macchine prodotte.

    Regola corretta per il KPI:
    - conta 1 macchina per ogni chiusura_finale valida;
    - non somma QuantitaConforme;
    - non somma Quantita;
    - non conta chiusura_macchina, per evitare doppioni su ordini multifase.
    """
    azione = _norm_text(getattr(rt, "Azione", "")).lower()

    if azione != "chiusura_finale":
        return 0.0

    return 1.0


def _kpi_parse_date(value) -> date | None:
    raw = _norm_text(value)
    if not raw:
        return None

    raw = raw[:19]

    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass

    return None


def _snapshot_int(value) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _snapshot_row_to_dict(row: ProductionKpiSnapshot) -> dict:
    return {
        "id": row.id,
        "snapshot_month": row.snapshot_month,
        "scope_type": row.scope_type,
        "scope_code": row.scope_code,
        "period_start": row.period_start,
        "period_end": row.period_end,
        "ordini_chiusi": _snapshot_int(row.ordini_chiusi),
        "ordini_in_ritardo": _snapshot_int(row.ordini_in_ritardo),
        "macchine_prodotte": round(
            _snapshot_float(getattr(row, "macchine_prodotte", 0)), 2
        ),
        "percentuale_ritardo": round(_snapshot_float(row.percentuale_ritardo), 2),
        "giorni_medi_ritardo": round(_snapshot_float(row.giorni_medi_ritardo), 2),
        "tempo_previsto_totale": round(_snapshot_float(row.tempo_previsto_totale), 2),
        "tempo_reale_totale": round(_snapshot_float(row.tempo_reale_totale), 2),
        "scostamento_totale": round(_snapshot_float(row.scostamento_totale), 2),
        "scostamento_percentuale": round(
            _snapshot_float(row.scostamento_percentuale), 2
        ),
        "tempo_medio_ordine": round(_snapshot_float(row.tempo_medio_ordine), 2),
        "tempo_medio_fase": round(_snapshot_float(row.tempo_medio_fase), 2),
        "created_at": row.created_at or "",
        "created_by": row.created_by or "",
    }


def _home_config_int(value, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _autosize_worksheet(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)


def _write_sheet_from_rows(ws, headers: list[tuple[str, str]], rows: list[dict]):
    for col_idx, (_, label) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows or [], start=2):
        for col_idx, (key, _) in enumerate(headers, start=1):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=_excel_safe(row.get(key)),
            )

    ws.freeze_panes = "A2"
    _autosize_worksheet(ws)


def _add_aggregate_sheet(wb, title: str, rows: list[dict]):
    ws = wb.create_sheet(title=title)

    headers = [
        ("key", "Voce"),
        ("ordini", "Ordini"),
        ("ritardi", "Ordini in ritardo"),
        ("percentuale_ritardo", "% ritardo"),
        ("giorni_medi_ritardo", "Giorni medi ritardo"),
        ("tempo_previsto", "Tempo previsto"),
        ("tempo_reale", "Tempo reale"),
        ("scostamento", "Scostamento"),
        ("scostamento_percentuale", "% scostamento"),
    ]

    _write_sheet_from_rows(ws, headers, rows)


def _kpi_event_is_eligible(rt: OdpRuntimeLog, il: InputOdpLog | None = None) -> bool:
    azione = _norm_text(getattr(rt, "Azione", "")).lower()
    topic = _norm_text(getattr(rt, "Topic", "")).lower()
    motivo = _norm_text(getattr(rt, "Motivo", "")).lower()
    payload = _norm_text(getattr(rt, "PayloadJson", "")).lower()

    if azione not in {"chiusura_finale", "chiusura_macchina"}:
        return False

    if "eliminato_gestionale" in azione:
        return False

    if "eliminato_gestionale" in topic:
        return False

    if "eliminato dal gestionale" in motivo:
        return False

    if "eliminato_gestionale" in payload:
        return False

    if il is not None:
        chiusura_parziale = _norm_text(getattr(il, "ChiusuraParziale", "")).lower()
        if chiusura_parziale in {"1", "true", "si", "sì", "yes"}:
            return False

    return True


def _kpi_jsonish_list(value) -> list[str]:
    raw = _norm_text(value)
    if not raw:
        return []

    try:
        parsed = json.loads(raw)
    except Exception:
        return [raw]

    if not isinstance(parsed, list):
        parsed = [parsed]

    out = []
    for item in parsed:
        value = _norm_text(item)
        if value:
            out.append(value)

    return out


def _kpi_active_value_from_list(raw_values, raw_phases, fase: str) -> str:
    values = _kpi_jsonish_list(raw_values)
    phases = _parse_phase_list(raw_phases)
    fase = _norm_text(fase)

    if not values:
        return ""

    if phases and len(phases) == len(values):
        for phase, value in zip(phases, values):
            if _norm_text(phase) == fase:
                return _norm_text(value)

    fase_int = _fase_to_int(fase)
    if fase_int is not None:
        idx = fase_int - 1
        if 0 <= idx < len(values):
            return _norm_text(values[idx])

    if len(values) == 1:
        return _norm_text(values[0])

    return ""


def _kpi_reparto_for_log(rt: OdpRuntimeLog, il: InputOdpLog | None) -> str:
    if il is not None:
        fase = _norm_text(getattr(il, "FaseConsuntivata", "")) or _norm_text(
            getattr(il, "FaseAttiva", "")
        )
        value = _kpi_active_value_from_list(
            getattr(il, "CodReparto", ""),
            getattr(il, "NumFase", ""),
            fase,
        )
        return _first_code_from_cell(value) or _first_code_from_cell(
            getattr(il, "CodReparto", "")
        )

    return _first_code_from_cell(getattr(rt, "CodReparto", ""))


def _kpi_risorsa_for_log(il: InputOdpLog | None) -> str:
    if il is None:
        return ""

    if _norm_text(getattr(il, "RisorsaAttiva", "")):
        return _norm_text(getattr(il, "RisorsaAttiva", ""))

    fase = _norm_text(getattr(il, "FaseConsuntivata", "")) or _norm_text(
        getattr(il, "FaseAttiva", "")
    )

    value = _kpi_active_value_from_list(
        getattr(il, "CodRisorsaProd", ""),
        getattr(il, "NumFase", ""),
        fase,
    )

    return _first_code_from_cell(value)


def _kpi_lavorazione_for_log(il: InputOdpLog | None) -> str:
    if il is None:
        return ""

    if _norm_text(getattr(il, "LavorazioneAttiva", "")):
        return _norm_text(getattr(il, "LavorazioneAttiva", ""))

    fase = _norm_text(getattr(il, "FaseConsuntivata", "")) or _norm_text(
        getattr(il, "FaseAttiva", "")
    )

    value = _kpi_active_value_from_list(
        getattr(il, "CodLavorazione", ""),
        getattr(il, "NumFase", ""),
        fase,
    )

    return _first_code_from_cell(value)


def _kpi_matches_filters(row: dict, filters: dict) -> bool:
    for key in ("reparto", "risorsa", "lavorazione", "operatore", "articolo", "stato"):
        expected = _norm_text(filters.get(key)).lower()
        if not expected:
            continue

        current = _norm_text(row.get(key)).lower()

        if expected not in current:
            return False

    return True


def _apply_kpi_group(bucket: dict, row: dict) -> None:
    bucket["ordini"] += 1
    bucket["tempo_previsto"] += float(row.get("tempo_previsto_ore") or 0.0)
    bucket["tempo_reale"] += float(row.get("tempo_reale_ore") or 0.0)
    bucket["scostamento"] += float(row.get("scostamento_ore") or 0.0)

    ritardo_giorni = float(row.get("ritardo_giorni") or 0.0)
    if ritardo_giorni > 0:
        bucket["ritardi"] += 1
        bucket["giorni_ritardo_totali"] += ritardo_giorni


def _finalize_kpi_group(bucket: dict) -> dict:
    ordini = int(bucket.get("ordini") or 0)
    ritardi = int(bucket.get("ritardi") or 0)

    tempo_previsto = float(bucket.get("tempo_previsto") or 0.0)
    tempo_reale = float(bucket.get("tempo_reale") or 0.0)
    scostamento = float(bucket.get("scostamento") or 0.0)

    return {
        "key": bucket["key"],
        "ordini": ordini,
        "ritardi": ritardi,
        "percentuale_ritardo": round((ritardi / ordini) * 100, 2) if ordini else 0.0,
        "giorni_medi_ritardo": round((bucket["giorni_ritardo_totali"] / ritardi), 2)
        if ritardi
        else 0.0,
        "tempo_previsto": round(tempo_previsto, 2),
        "tempo_reale": round(tempo_reale, 2),
        "scostamento": round(scostamento, 2),
        "scostamento_percentuale": round((scostamento / tempo_previsto) * 100, 2)
        if tempo_previsto > 0
        else 0.0,
    }


def _dashboard_produzione_default_section(policy: RbacPolicy) -> str:
    if policy.can("dashboard_produzione"):
        return "cruscotto"

    if policy.can("kpi_produzione"):
        return "kpi"

    return ""


def _dashboard_stato_norm(ordine: InputOdp) -> str:
    runtime = getattr(ordine, "runtime_row", None)

    stato = (
        _norm_text(getattr(runtime, "Stato_odp", ""))
        or _norm_text(getattr(ordine, "StatoOrdine", ""))
        or _norm_text(getattr(ordine, "StatoOrdineErp", ""))
    )

    return stato.strip()


def _dashboard_fase_attiva(ordine: InputOdp) -> str:
    runtime = getattr(ordine, "runtime_row", None)

    return (
        _norm_text(getattr(runtime, "FaseAttiva", ""))
        or _norm_text(getattr(ordine, "FaseAttiva", ""))
        or "1"
    )


def _dashboard_reparti_label_map() -> dict[str, str]:
    rows = Reparti.query.order_by(Reparti.Codice.asc()).all()
    out: dict[str, str] = {}

    for row in rows:
        codice = _norm_text(getattr(row, "Codice", ""))
        descrizione = _norm_text(getattr(row, "Descrizione", ""))

        if not codice:
            continue

        if descrizione and descrizione.lower() != codice.lower():
            out[codice] = f"{descrizione} ({codice})"
        else:
            out[codice] = codice

    return out


def _dashboard_active_value(ordine: InputOdp, attr_name: str) -> str:
    fase_attiva = _dashboard_fase_attiva(ordine)

    return _active_value_for_phase(
        getattr(ordine, attr_name, ""),
        getattr(ordine, "NumFase", ""),
        fase_attiva,
    )


def _dashboard_reparto_attivo(ordine: InputOdp) -> str:
    raw = _dashboard_active_value(ordine, "CodReparto")
    return _first_code_from_cell(raw) or _first_code_from_cell(
        getattr(ordine, "CodReparto", "")
    )


def _dashboard_risorsa_attiva(ordine: InputOdp) -> str:
    runtime = getattr(ordine, "runtime_row", None)

    return (
        _norm_text(getattr(runtime, "RisorsaAttiva", ""))
        or _norm_text(getattr(ordine, "RisorsaAttiva", ""))
        or _first_code_from_cell(_dashboard_active_value(ordine, "CodRisorsaProd"))
    )


def _dashboard_data_fine_prevista(ordine: InputOdp):
    raw_data_fine = getattr(ordine, "DataFinePrevista", "") or getattr(
        ordine, "DataFineSched", ""
    )

    fase_attiva = getattr(ordine, "FaseAttiva", "1")
    data_fine_attiva = InputOdp._active_value_from_phase_list(
        raw_data_fine,
        fase_attiva,
    )

    return _dashboard_parse_date(data_fine_attiva)


def _dashboard_parse_date(value) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    raw = _norm_text(value)

    if not raw:
        return None

    if raw.lower() in {"none", "null", "nan", "nat", "0000-00-00"}:
        return None

    # Gestione valori salvati come lista JSON:
    # es. ["2026-04-07 00:00:00"]
    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list) and parsed:
                raw = _norm_text(parsed[0])
            else:
                return None
        except (json.JSONDecodeError, TypeError, ValueError):
            return None

    # Gestione valori JSON string:
    # es. "2026-04-07 00:00:00"
    elif raw.startswith('"') and raw.endswith('"'):
        try:
            raw = _norm_text(json.loads(raw))
        except (json.JSONDecodeError, TypeError, ValueError):
            raw = raw.strip('"')

    iso_raw = raw.replace("Z", "+00:00")

    try:
        return datetime.fromisoformat(iso_raw).date()
    except ValueError:
        pass

    clean = raw.replace("T", " ").strip()
    clean = re.sub(r"\.\d+", "", clean)
    clean = re.sub(r"\s*(Z|[+-]\d{2}:?\d{2})$", "", clean).strip()

    formats = (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%Y%m%d",
        "%d%m%Y",
    )

    for fmt in formats:
        try:
            return datetime.strptime(clean, fmt).date()
        except ValueError:
            continue

    return None


def _dashboard_tempo_previsto_ore(ordine: InputOdp) -> float:
    """
    Restituisce le ore previste della fase attiva.

    Gestisce anche valori salvati come lista JSON, ad esempio:
    ["1.5", "2.0", "0.75"]
    """

    raw = _dashboard_active_value(ordine, "TempoPrevistoLavoraz")

    if not raw:
        raw = getattr(ordine, "TempoPrevistoLavoraz", "")

    return _safe_float(raw)


def _dashboard_order_key(ordine: InputOdp) -> str:
    return f"{_norm_text(ordine.IdDocumento)}|{_norm_text(ordine.IdRiga)}|{_dashboard_fase_attiva(ordine)}"


def _dashboard_order_payload(ordine: InputOdp, tipo_criticita: str = "") -> dict:
    runtime = getattr(ordine, "runtime_row", None)
    data_fine = _dashboard_data_fine_prevista(ordine)
    today = _dashboard_today()
    ritardo_giorni = 0

    if data_fine and data_fine < today:
        ritardo_giorni = (today - data_fine).days

    return {
        "key": _dashboard_order_key(ordine),
        "tipo": tipo_criticita,
        "ordine": _dashboard_order_label(ordine),
        "cod_art": _norm_text(getattr(ordine, "CodArt", "")),
        "descrizione": _norm_text(getattr(ordine, "DesArt", "")),
        "reparto": _dashboard_reparto_attivo(ordine),
        "risorsa": _dashboard_risorsa_attiva(ordine),
        "lavorazione": _dashboard_lavorazione_attiva(ordine),
        "stato": _dashboard_stato_norm(ordine),
        "fase": _dashboard_fase_attiva(ordine),
        "operatore": _norm_text(getattr(runtime, "Utente_operazione", ""))
        if runtime
        else "",
        "data_fine_prevista": data_fine.isoformat() if data_fine else "",
        "ritardo_giorni": ritardo_giorni,
        "tempo_previsto_ore": _dashboard_carico_ore(ordine),
        "priorita": "",
    }


def _dashboard_next_month_days() -> list[date]:
    today = _dashboard_today()
    return [today + timedelta(days=i) for i in range(DASHBOARD_PRODUZIONE_FUTURE_DAYS)]


def _dashboard_capacity_hours_for_next_days(
    *,
    scope_type: str = "global",
    scope_code: str = "*",
) -> float:
    capacity = _dashboard_capacity_by_weekday(
        scope_type=scope_type,
        scope_code=scope_code,
    )
    return round(
        sum(
            float(capacity.get(day.weekday(), 0.0) or 0.0)
            for day in _dashboard_next_month_days()
        ),
        2,
    )


def _dashboard_carico_per_reparto(ordini: list[InputOdp]) -> list[dict]:
    buckets = {}
    reparto_labels = _dashboard_reparti_label_map()

    for ordine in ordini or []:
        stato = _dashboard_stato_norm(ordine).lower()
        reparto = _dashboard_reparto_attivo(ordine) or "-"
        reparto_label = _dashboard_reparto_label(reparto, reparto_labels)
        ore = _dashboard_carico_ore(ordine)

        buckets.setdefault(
            reparto,
            {
                "label": reparto_label,
                "reparto": reparto,
                "codice_reparto": reparto,
                "ore_attive": 0.0,
                "ore_sospese": 0.0,
                "ore_pianificate": 0.0,
                "ore_totali": 0.0,
            },
        )

        if "attiv" in stato:
            buckets[reparto]["ore_attive"] += ore
        elif "sospes" in stato:
            buckets[reparto]["ore_sospese"] += ore
        elif "pianificat" in stato:
            buckets[reparto]["ore_pianificate"] += ore

    out = []
    for row in buckets.values():
        row["ore_attive"] = round(float(row["ore_attive"] or 0.0), 2)
        row["ore_sospese"] = round(float(row["ore_sospese"] or 0.0), 2)
        row["ore_pianificate"] = round(float(row["ore_pianificate"] or 0.0), 2)
        row["ore_totali"] = round(
            row["ore_attive"] + row["ore_sospese"] + row["ore_pianificate"],
            2,
        )
        out.append(row)

    return sorted(out, key=lambda x: (-x["ore_totali"], x["label"].lower()))[:12]


def _dashboard_carico_per_risorsa_chart(carico_rows: list[dict]) -> list[dict]:
    out = []
    for row in carico_rows or []:
        out.append(
            {
                "label": row.get("risorsa") or "-",
                "ore_attive": round(float(row.get("ore_attive") or 0.0), 2),
                "ore_sospese": round(float(row.get("ore_sospese") or 0.0), 2),
                "ore_pianificate": round(float(row.get("ore_pianificate") or 0.0), 2),
                "ore_totali": round(float(row.get("ore_totali") or 0.0), 2),
            }
        )

    return sorted(out, key=lambda x: (-x["ore_totali"], x["label"].lower()))[:12]


def _dashboard_cruscotto_empty_payload() -> dict:
    return {
        "cards": {
            "ordini_attivi": 0,
            "ordini_sospesi": 0,
            "ordini_pianificati": 0,
            "tempo_previsto_residuo": 0.0,
            "operatori_impegnati": 0,
            "operatori_capacita": 0,
            "ordini_critici": 0,
            "risorse_sovraccariche": 0,
            "ordini_in_ritardo": 0,
            "ordini_scadenza_oggi": 0,
            "ordini_senza_tempo_previsto": 0,
            "ordini_collaudo": 0,
        },
        "charts": {
            "carico_prossimi_giorni": [],
            "stati_ordine": [],
            "carico_per_risorsa": [],
            "carico_per_reparto": [],
            "saturazione_risorse": [],
        },
        "criticita": [],
        "details": [],
        "operatori": [],
        "carico_risorsa": [],
        "collaudo": [],
        "capacity_calendar": [],
        "filters": {},
        "filter_options": _dashboard_empty_filter_options(),
    }


def _dashboard_capacity_users(
    policy: RbacPolicy, filters: dict | None = None
) -> list[User]:
    """
    Restituisce gli operatori attivi da considerare nel calcolo capacità.

    Regole:
    - considera solo utenti active = True;
    - rispetta i reparti consentiti dalla policy;
    - applica i filtri cruscotto;
    - esclude utenti senza RepartoPrinc, salvo override operatore esplicito;
    - include un utente senza reparto solo se ha capacità specifica operatore.
    """

    filters = filters or {}

    allowed_reparti = {
        _norm_text(code)
        for code in getattr(policy, "allowed_reparti", [])
        if _norm_text(code)
    }

    users = User.query.filter(User.active.is_(True)).order_by(User.username.asc()).all()

    out = []

    for user in users:
        user_reparto = _norm_text(getattr(user, "RepartoPrinc", ""))
        operator_code = str(int(user.id))

        has_operator_capacity = _dashboard_capacity_rows_exist(
            "operatore",
            operator_code,
        )

        # Esclude admin, utenti tecnici o generici senza reparto,
        # a meno che abbiano una capacità operatore configurata.
        if not user_reparto and not has_operator_capacity:
            continue

        # Rispetta i reparti consentiti dalla policy.
        # Se l'utente non ha reparto ma ha override operatore, passa.
        if allowed_reparti and user_reparto and user_reparto not in allowed_reparti:
            continue

        if not _dashboard_user_matches_capacity_filters(user, filters):
            continue

        out.append(user)

    return out


def _dashboard_capacity_rows_exist(scope_type: str, scope_code: str) -> bool:
    scope_type = _norm_text(scope_type)
    scope_code = _norm_text(scope_code)

    if not scope_type or not scope_code:
        return False

    return (
        ProductionCapacityCalendar.query.filter_by(
            scope_type=scope_type,
            scope_code=scope_code,
        ).first()
        is not None
    )


def _dashboard_capacity_by_weekday_for_policy(
    policy: RbacPolicy,
    filters: dict | None = None,
) -> tuple[dict[int, float], int]:
    """
    Calcola la capacità produttiva giornaliera totale per il cruscotto.

    Logica:
    - prende gli operatori attivi coerenti con policy e filtri;
    - per ogni operatore calcola la capacità settimanale con priorità:
        1. capacità specifica operatore;
        2. capacità reparto;
        3. capacità globale;
    - somma le ore per ogni giorno della settimana.

    Ritorna:
        (
            {
                0: ore_lunedì,
                1: ore_martedì,
                ...
                6: ore_domenica,
            },
            numero_operatori_considerati
        )
    """

    filters = filters or {}

    totals = {weekday: 0.0 for weekday in range(7)}
    operator_count = 0

    users = _dashboard_capacity_users(policy, filters)

    for user in users:
        capacity = _dashboard_capacity_for_operator(user)

        # Conta l'operatore solo se ha almeno una capacità settimanale valorizzata.
        # Se vuoi contare anche operatori con 0 ore su tutta la settimana,
        # rimuovi questo controllo.
        has_capacity = any(
            float(capacity.get(weekday, 0.0) or 0.0) > 0 for weekday in range(7)
        )

        if not has_capacity:
            continue

        operator_count += 1

        for weekday in range(7):
            totals[weekday] += float(capacity.get(weekday, 0.0) or 0.0)

    return (
        {weekday: round(hours, 2) for weekday, hours in totals.items()},
        operator_count,
    )


def _dashboard_seed_user_filter_options(options: dict) -> None:
    """
    Alimenta il filtro Operatore da tutti gli utenti presenti in users.
    Non filtra per active, ruolo, reparto o policy.
    """
    users = User.query.order_by(func.lower(User.username)).all()

    for user in users:
        username = _norm_text(getattr(user, "username", ""))

        if not username:
            continue

        _dashboard_add_filter_option(
            options,
            "operatore",
            username,
            username,
        )


def _dashboard_seed_master_filter_options(options: dict) -> None:
    """
    Filtri caricati da anagrafiche/models, indipendenti dalle righe visibili.
    """
    _dashboard_seed_model_filter_options(
        options,
        key="reparto",
        model=Reparti,
    )

    _dashboard_seed_model_filter_options(
        options,
        key="risorsa",
        model=Risorse,
    )

    _dashboard_seed_model_filter_options(
        options,
        key="lavorazione",
        model=Lavorazioni,
    )

    _dashboard_seed_user_filter_options(options)
    _dashboard_seed_stato_filter_options(options)


def _kpi_parse_datetime(value) -> datetime | None:
    raw = _norm_text(value)
    if not raw:
        return None

    try:
        dt = datetime.fromisoformat(raw)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ROME_TZ)
        return dt.astimezone(ROME_TZ)
    except Exception:
        pass

    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(raw[:19], fmt)
            return dt.replace(tzinfo=ROME_TZ)
        except ValueError:
            pass

    return None


def _kpi_date_range_from_request() -> tuple[date, date]:
    today = _dashboard_today()

    default_from = today - timedelta(days=365)
    default_to = today

    date_from = _kpi_parse_date(request.args.get("date_from")) or default_from
    date_to = _kpi_parse_date(request.args.get("date_to")) or default_to

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    return date_from, date_to


def _kpi_tempo_previsto_ore(il: InputOdpLog | None) -> float:
    if il is None:
        return 0.0

    fase = _norm_text(getattr(il, "FaseConsuntivata", "")) or _norm_text(
        getattr(il, "FaseAttiva", "")
    )

    raw = _kpi_active_value_from_list(
        getattr(il, "TempoPrevistoLavoraz", ""),
        getattr(il, "NumFase", ""),
        fase,
    )

    value = _safe_float(raw)

    if value > 0:
        return value

    return _safe_float(getattr(il, "TempoPrevistoLavoraz", ""))


def _kpi_tempo_reale_ore(rt: OdpRuntimeLog, il: InputOdpLog | None) -> float:
    if il is not None:
        value = _safe_float(getattr(il, "TempoFunzionamentoFinale", ""))
        if value > 0:
            return value

    value = _safe_float(getattr(rt, "TempoFunzionamentoPost", ""))
    if value > 0:
        return value

    elapsed_seconds = _safe_float(getattr(rt, "ElapsedSeconds", ""))
    if elapsed_seconds > 0:
        return elapsed_seconds / 3600.0

    return 0.0


def _kpi_closed_at(rt: OdpRuntimeLog, il: InputOdpLog | None) -> datetime | None:
    if il is not None:
        dt = _kpi_parse_datetime(getattr(il, "ClosedAt", ""))
        if dt is not None:
            return dt

    return _kpi_parse_datetime(getattr(rt, "EventAt", ""))


def _kpi_data_fine_prevista(il: InputOdpLog | None) -> date | None:
    if il is None:
        return None

    return _kpi_parse_date(getattr(il, "DataFineSched", ""))


def _kpi_is_collaudo(reparto: str, risorsa: str, lavorazione: str) -> bool:
    return (
        _norm_text(reparto) == "70"
        or "coll" in _norm_text(risorsa).lower()
        or "coll" in _norm_text(lavorazione).lower()
    )


def _new_kpi_group_bucket(key: str) -> dict:
    return {
        "key": key,
        "ordini": 0,
        "ritardi": 0,
        "giorni_ritardo_totali": 0.0,
        "tempo_previsto": 0.0,
        "tempo_reale": 0.0,
        "scostamento": 0.0,
    }


def _build_dashboard_kpi_payload(*, detail_limit: int | None = 500) -> dict:
    payload = _dashboard_kpi_empty_payload()
    date_from, date_to = _kpi_date_range_from_request()
    date_from_dt = datetime.combine(date_from, datetime.min.time()).replace(
        tzinfo=ROME_TZ
    )
    date_to_dt = datetime.combine(date_to, datetime.max.time()).replace(tzinfo=ROME_TZ)

    filters = {
        "reparto": request.args.get("reparto"),
        "risorsa": request.args.get("risorsa"),
        "lavorazione": request.args.get("lavorazione"),
        "operatore": request.args.get("operatore"),
        "articolo": request.args.get("articolo"),
        "stato": request.args.get("stato"),
    }

    rows = (
        db.session.query(OdpRuntimeLog, InputOdpLog)
        .outerjoin(
            InputOdpLog,
            and_(
                InputOdpLog.OperationGroupId == OdpRuntimeLog.OperationGroupId,
                InputOdpLog.IdDocumento == OdpRuntimeLog.IdDocumento,
                InputOdpLog.IdRiga == OdpRuntimeLog.IdRiga,
            ),
        )
        .filter(OdpRuntimeLog.Azione.in_(["chiusura_finale", "chiusura_macchina"]))
        .order_by(OdpRuntimeLog.EventAt.desc(), OdpRuntimeLog.log_id.desc())
        .all()
    )

    detail_rows = []
    today = _dashboard_today()
    filter_options = _dashboard_new_filter_options_bucket()
    filter_label_maps = _dashboard_filter_label_maps()
    _dashboard_seed_master_filter_options(filter_options)

    groups = {
        "reparti": {},
        "risorse": {},
        "lavorazioni": {},
        "operatori": {},
        "articoli": {},
    }

    collaudo_tempi = []
    collaudi_chiusi_oggi = 0
    for rt, il in rows:
        if not _kpi_event_is_eligible(rt, il):
            continue

        closed_at = _kpi_closed_at(rt, il)
        if closed_at is None:
            continue

        if closed_at < date_from_dt or closed_at > date_to_dt:
            continue

        reparto = _kpi_reparto_for_log(rt, il)
        risorsa = _kpi_risorsa_for_log(il)
        lavorazione = _kpi_lavorazione_for_log(il)
        macchine_prodotte = _kpi_macchine_prodotte(rt, il)

        tempo_previsto = _kpi_tempo_previsto_ore(il)
        tempo_reale = _kpi_tempo_reale_ore(rt, il)
        scostamento = tempo_reale - tempo_previsto

        data_fine_prevista = _kpi_data_fine_prevista(il)
        ritardo_giorni = 0

        if data_fine_prevista and closed_at.date() > data_fine_prevista:
            ritardo_giorni = (closed_at.date() - data_fine_prevista).days

        articolo = (
            _norm_text(getattr(il, "CodArt", ""))
            if il
            else _norm_text(getattr(rt, "CodArt", ""))
        )
        descrizione = _norm_text(getattr(il, "DesArt", "")) if il else ""
        operatore = _norm_text(getattr(rt, "UtenteOperazione", ""))
        stato = _norm_text(getattr(rt, "StatoOrdinePost", "")) or _norm_text(
            getattr(rt, "StatoOdpPost", "")
        )

        row = {
            "operation_group_id": _norm_text(getattr(rt, "OperationGroupId", "")),
            "id_documento": _norm_text(getattr(rt, "IdDocumento", "")),
            "id_riga": _norm_text(getattr(rt, "IdRiga", "")),
            "rif_registraz": _norm_text(getattr(rt, "RifRegistraz", "")),
            "ordine": f"{_norm_text(getattr(rt, 'IdDocumento', ''))}/{_norm_text(getattr(rt, 'IdRiga', ''))}",
            "event_at": closed_at.isoformat(timespec="seconds"),
            "data_chiusura": closed_at.date().isoformat(),
            "azione": _norm_text(getattr(rt, "Azione", "")),
            "reparto": reparto,
            "risorsa": risorsa,
            "lavorazione": lavorazione,
            "operatore": operatore,
            "articolo": articolo,
            "descrizione": descrizione,
            "stato": stato,
            "fase": _norm_text(getattr(il, "FaseConsuntivata", ""))
            if il
            else _norm_text(getattr(rt, "FasePost", "")),
            "tempo_previsto_ore": round(tempo_previsto, 2),
            "tempo_reale_ore": round(tempo_reale, 2),
            "macchine_prodotte": round(macchine_prodotte, 2),
            "scostamento_ore": round(scostamento, 2),
            "scostamento_percentuale": round((scostamento / tempo_previsto) * 100, 2)
            if tempo_previsto > 0
            else None,
            "data_fine_prevista": data_fine_prevista.isoformat()
            if data_fine_prevista
            else "",
            "ritardo_giorni": ritardo_giorni,
            "is_ritardo": ritardo_giorni > 0,
            "is_collaudo": _kpi_is_collaudo(reparto, risorsa, lavorazione),
        }

        _dashboard_collect_filter_options_from_row(
            filter_options,
            {
                **row,
                "cod_art": row.get("articolo"),
            },
            filter_label_maps,
        )

        if not _kpi_matches_filters(row, filters):
            continue

        detail_rows.append(row)

        for group_name, group_key in (
            ("reparti", reparto or "-"),
            ("risorse", risorsa or "-"),
            ("lavorazioni", lavorazione or "-"),
            ("operatori", operatore or "-"),
            ("articoli", articolo or "-"),
        ):
            groups[group_name].setdefault(group_key, _new_kpi_group_bucket(group_key))
            _apply_kpi_group(groups[group_name][group_key], row)

        if row["is_collaudo"]:
            collaudo_tempi.append(tempo_reale)
            if closed_at.date() == today:
                collaudi_chiusi_oggi += 1

    ordini = len(detail_rows)
    ritardi = sum(1 for row in detail_rows if row["is_ritardo"])
    macchine_prodotte = sum(
        float(row.get("macchine_prodotte", 0.0) or 0.0) for row in detail_rows
    )

    tempo_previsto_totale = sum(
        float(row["tempo_previsto_ore"] or 0.0) for row in detail_rows
    )
    tempo_reale_totale = sum(
        float(row["tempo_reale_ore"] or 0.0) for row in detail_rows
    )
    scostamento_totale = tempo_reale_totale - tempo_previsto_totale

    giorni_ritardo_totali = sum(
        float(row["ritardo_giorni"] or 0.0) for row in detail_rows if row["is_ritardo"]
    )

    rows_con_tempo_previsto = [
        row for row in detail_rows if float(row.get("tempo_previsto_ore") or 0.0) > 0
    ]
    ordini_tempi_affidabili = sum(
        1
        for row in rows_con_tempo_previsto
        if float(row.get("tempo_reale_ore") or 0.0)
        <= float(row.get("tempo_previsto_ore") or 0.0) * 1.10
    )
    ordini_tempi_non_affidabili = max(
        len(rows_con_tempo_previsto) - ordini_tempi_affidabili,
        0,
    )
    affidabilita_tempi = (
        round((ordini_tempi_affidabili / len(rows_con_tempo_previsto)) * 100, 2)
        if rows_con_tempo_previsto
        else 0.0
    )

    payload["cards"] = {
        "ordini_chiusi": ordini,
        "ordini_in_ritardo": ritardi,
        "percentuale_ritardo": round((ritardi / ordini) * 100, 2) if ordini else 0.0,
        "giorni_medi_ritardo": round(giorni_ritardo_totali / ritardi, 2)
        if ritardi
        else 0.0,
        "macchine_prodotte": round(macchine_prodotte, 2),
        "tempo_previsto_totale": round(tempo_previsto_totale, 2),
        "tempo_reale_totale": round(tempo_reale_totale, 2),
        "tempo_reale_vs_previsto": round(tempo_reale_totale - tempo_previsto_totale, 2),
        "scostamento_totale": round(scostamento_totale, 2),
        "scostamento_percentuale": round(
            (scostamento_totale / tempo_previsto_totale) * 100, 2
        )
        if tempo_previsto_totale > 0
        else 0.0,
        "scostamento_medio": round(scostamento_totale / ordini, 2) if ordini else 0.0,
        "tempo_medio_ordine": round(tempo_reale_totale / ordini, 2) if ordini else 0.0,
        "tempo_medio_fase": round(tempo_reale_totale / ordini, 2) if ordini else 0.0,
        "tempo_medio_collaudo": round(sum(collaudo_tempi) / len(collaudo_tempi), 2)
        if collaudo_tempi
        else 0.0,
        "collaudi_chiusi_oggi": collaudi_chiusi_oggi,
        "affidabilita_tempi": affidabilita_tempi,
        "ordini_tempi_affidabili": ordini_tempi_affidabili,
        "ordini_tempi_non_affidabili": ordini_tempi_non_affidabili,
    }

    payload["charts"]["tempo_reale_vs_previsto"] = [
        {
            "label": "Previsto",
            "value": round(tempo_previsto_totale, 2),
        },
        {
            "label": "Reale",
            "value": round(tempo_reale_totale, 2),
        },
    ]

    payload["aggregati"] = {
        name: sorted(
            [_finalize_kpi_group(bucket) for bucket in group.values()],
            key=lambda x: (-x["ordini"], x["key"].lower()),
        )[:50]
        for name, group in groups.items()
    }

    reparto_labels = _dashboard_reparti_label_map()

    payload["charts"]["ritardo_medio_reparto"] = [
        {
            "label": _dashboard_reparto_label(row["key"], reparto_labels),
            "codice_reparto": row["key"],
            "value": row["giorni_medi_ritardo"],
            "ordini": row["ordini"],
            "ritardi": row["ritardi"],
        }
        for row in sorted(
            payload["aggregati"]["reparti"],
            key=lambda x: (
                -float(x.get("giorni_medi_ritardo") or 0.0),
                -int(x.get("ritardi") or 0),
                -int(x.get("ordini") or 0),
                x.get("key", "").lower(),
            ),
        )
        if int(row.get("ordini") or 0) > 0
    ][:10]

    payload["charts"]["ritardo_medio_risorsa"] = [
        {
            "label": row["key"],
            "value": row["giorni_medi_ritardo"],
        }
        for row in payload["aggregati"]["risorse"]
        if row["giorni_medi_ritardo"] > 0
    ]

    payload["charts"]["top_risorse_scostamento"] = [
        {
            "label": row["key"],
            "value": row["scostamento_percentuale"],
            "scostamento_ore": row["scostamento"],
        }
        for row in sorted(
            payload["aggregati"]["risorse"],
            key=lambda x: (
                -float(x.get("scostamento_percentuale") or 0.0),
                -float(x.get("scostamento") or 0.0),
                x.get("key", "").lower(),
            ),
        )
        if float(row.get("scostamento_percentuale") or 0.0) > 0
    ][:10]

    payload["charts"]["top_reparti_scostamento"] = [
        {
            "label": _dashboard_reparto_label(row["key"], reparto_labels),
            "codice_reparto": row["key"],
            "value": row["scostamento_percentuale"],
            "scostamento_ore": row["scostamento"],
        }
        for row in sorted(
            payload["aggregati"]["reparti"],
            key=lambda x: (
                -float(x.get("scostamento_percentuale") or 0.0),
                -float(x.get("scostamento") or 0.0),
                x.get("key", "").lower(),
            ),
        )
        if float(row.get("scostamento_percentuale") or 0.0) > 0
    ][:10]

    payload["charts"]["top_lavorazioni_scostamento"] = [
        {
            "label": row["key"],
            "value": row["scostamento_percentuale"],
            "scostamento_ore": row["scostamento"],
        }
        for row in sorted(
            payload["aggregati"]["lavorazioni"],
            key=lambda x: (
                -float(x.get("scostamento_percentuale") or 0.0),
                -float(x.get("scostamento") or 0.0),
                x.get("key", "").lower(),
            ),
        )
        if float(row.get("scostamento_percentuale") or 0.0) > 0
    ][:10]

    payload["charts"]["affidabilita_tempi"] = [
        {"label": "Entro +10%", "value": ordini_tempi_affidabili},
        {"label": "Oltre +10%", "value": ordini_tempi_non_affidabili},
    ]

    sorted_details = sorted(
        detail_rows,
        key=lambda x: x["event_at"],
        reverse=True,
    )

    if detail_limit is not None:
        sorted_details = sorted_details[:detail_limit]

    payload["details"] = sorted_details

    payload["filters"] = {
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        **{k: _norm_text(v) for k, v in filters.items()},
    }
    payload["filter_options"] = _dashboard_finalize_filter_options(filter_options)

    return payload


def _excel_safe(value):
    if value is None:
        return ""

    if isinstance(value, bool):
        return "Sì" if value else "No"

    return value


def _write_kpi_summary_sheet(ws, data: dict):
    ws.title = "Riepilogo"

    cards = data.get("cards") or {}
    filters = data.get("filters") or {}

    rows = [
        ("Periodo da", filters.get("date_from", "")),
        ("Periodo a", filters.get("date_to", "")),
        ("Filtro reparto", filters.get("reparto", "")),
        ("Filtro risorsa", filters.get("risorsa", "")),
        ("Filtro lavorazione", filters.get("lavorazione", "")),
        ("Filtro operatore", filters.get("operatore", "")),
        ("Filtro articolo", filters.get("articolo", "")),
        ("Filtro stato", filters.get("stato", "")),
        ("", ""),
        ("Ordini chiusi", cards.get("ordini_chiusi", 0)),
        ("Macchine prodotte", cards.get("macchine_prodotte", 0)),
        ("Ordini in ritardo", cards.get("ordini_in_ritardo", 0)),
        ("% ritardo", cards.get("percentuale_ritardo", 0)),
        ("Giorni medi ritardo", cards.get("giorni_medi_ritardo", 0)),
        ("Tempo previsto totale", cards.get("tempo_previsto_totale", 0)),
        ("Tempo reale totale", cards.get("tempo_reale_totale", 0)),
        ("Scostamento totale", cards.get("scostamento_totale", 0)),
        ("% scostamento", cards.get("scostamento_percentuale", 0)),
        ("Scostamento medio", cards.get("scostamento_medio", 0)),
        ("Tempo medio ordine", cards.get("tempo_medio_ordine", 0)),
        ("Tempo medio fase", cards.get("tempo_medio_fase", 0)),
        ("Tempo medio collaudo", cards.get("tempo_medio_collaudo", 0)),
        ("Collaudi chiusi oggi", cards.get("collaudi_chiusi_oggi", 0)),
    ]

    ws.cell(row=1, column=1, value="Parametro").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Valore").font = Font(bold=True)

    for idx, (label, value) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=_excel_safe(value))

    ws.freeze_panes = "A2"
    _autosize_worksheet(ws)


def _snapshot_float(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _snapshot_change(current_value, previous_value) -> dict:
    current_value = _snapshot_float(current_value)
    previous_value = _snapshot_float(previous_value)

    delta = current_value - previous_value

    if previous_value:
        delta_percent = (delta / previous_value) * 100
    else:
        delta_percent = 0.0

    return {
        "current": round(current_value, 2),
        "previous": round(previous_value, 2),
        "delta": round(delta, 2),
        "delta_percent": round(delta_percent, 2),
    }


def _build_kpi_snapshot_payload() -> dict:
    scope_type = _norm_text(request.args.get("scope_type")) or "global"
    scope_code = _norm_text(request.args.get("scope_code")) or "*"
    months_limit = _home_config_int(request.args.get("months"), 12)

    if months_limit <= 0:
        months_limit = 12

    if months_limit > 60:
        months_limit = 60

    query = ProductionKpiSnapshot.query

    if scope_type:
        query = query.filter(ProductionKpiSnapshot.scope_type == scope_type)

    if scope_code:
        query = query.filter(ProductionKpiSnapshot.scope_code == scope_code)

    rows = (
        query.order_by(ProductionKpiSnapshot.snapshot_month.desc())
        .limit(months_limit)
        .all()
    )

    rows_sorted = sorted(rows, key=lambda r: r.snapshot_month)

    data_rows = [_snapshot_row_to_dict(row) for row in rows_sorted]

    latest = data_rows[-1] if data_rows else None
    previous = data_rows[-2] if len(data_rows) >= 2 else None

    comparison = {}

    if latest:
        comparison = {
            "ordini_chiusi": _snapshot_change(
                latest.get("ordini_chiusi"),
                previous.get("ordini_chiusi") if previous else 0,
            ),
            "ordini_in_ritardo": _snapshot_change(
                latest.get("ordini_in_ritardo"),
                previous.get("ordini_in_ritardo") if previous else 0,
            ),
            "percentuale_ritardo": _snapshot_change(
                latest.get("percentuale_ritardo"),
                previous.get("percentuale_ritardo") if previous else 0,
            ),
            "tempo_reale_totale": _snapshot_change(
                latest.get("tempo_reale_totale"),
                previous.get("tempo_reale_totale") if previous else 0,
            ),
            "scostamento_totale": _snapshot_change(
                latest.get("scostamento_totale"),
                previous.get("scostamento_totale") if previous else 0,
            ),
            "macchine_prodotte": _snapshot_change(
                latest.get("macchine_prodotte"),
                previous.get("macchine_prodotte") if previous else 0,
            ),
        }

    available_scopes = (
        db.session.query(
            ProductionKpiSnapshot.scope_type,
            ProductionKpiSnapshot.scope_code,
        )
        .distinct()
        .order_by(
            ProductionKpiSnapshot.scope_type.asc(),
            ProductionKpiSnapshot.scope_code.asc(),
        )
        .all()
    )

    return {
        "filters": {
            "scope_type": scope_type,
            "scope_code": scope_code,
            "months": months_limit,
        },
        "available_scopes": [
            {
                "scope_type": st,
                "scope_code": sc,
                "label": f"{st}: {sc}",
            }
            for st, sc in available_scopes
        ],
        "latest": latest,
        "previous": previous,
        "comparison": comparison,
        "series": {
            "ordini_chiusi": [
                {
                    "month": row["snapshot_month"],
                    "value": row["ordini_chiusi"],
                }
                for row in data_rows
            ],
            "ordini_in_ritardo": [
                {
                    "month": row["snapshot_month"],
                    "value": row["ordini_in_ritardo"],
                }
                for row in data_rows
            ],
            "percentuale_ritardo": [
                {
                    "month": row["snapshot_month"],
                    "value": row["percentuale_ritardo"],
                }
                for row in data_rows
            ],
            "tempo_previsto_reale": [
                {
                    "month": row["snapshot_month"],
                    "tempo_previsto": row["tempo_previsto_totale"],
                    "tempo_reale": row["tempo_reale_totale"],
                    "scostamento": row["scostamento_totale"],
                }
                for row in data_rows
            ],
            "macchine_prodotte": [
                {
                    "month": row["snapshot_month"],
                    "value": row["macchine_prodotte"],
                }
                for row in data_rows
            ],
        },
        "rows": data_rows,
    }
